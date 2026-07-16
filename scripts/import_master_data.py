"""Import master data from templates/pvwood_master_data.xlsx into Postgres.

Idempotent (upsert by code). Auto-masters species. Resolves supplier codes and
reports unknown references instead of guessing. Run from repo root:

    python -m scripts.import_master_data [path-to-xlsx]
"""
from __future__ import annotations

import sys
from datetime import date, datetime

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.db import engine
from app.models.bom import GlueRecipe
from app.models.enums import ItemKind
from app.models.logs import Log, LogArrival
from app.models.master import Item, Species, WarehouseLocation
from app.models.procurement import Supplier
from app.models.sales import Customer
from scripts._import_utils import Report, intnum, num, s, sheet_rows

DEFAULT_PATH = "templates/pvwood_master_data.xlsx"
IN_TO_M = 0.0254
FT3_TO_M3 = 0.0283168


def to_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, (date, datetime)):
        return v.date() if isinstance(v, datetime) else v
    try:
        return datetime.strptime(s(v)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _cache(session: Session, model, key_attr: str) -> dict:
    return {getattr(o, key_attr): o for o in session.exec(select(model)).all()}


def run(path: str) -> None:
    wb = load_workbook(path, data_only=True)
    rep = Report()
    with Session(engine) as ses:
        locs = _cache(ses, WarehouseLocation, "code")
        sups = _cache(ses, Supplier, "code")
        custs = {c.name: c for c in ses.exec(select(Customer)).all()}
        items = _cache(ses, Item, "code")
        recipes = _cache(ses, GlueRecipe, "recipe_code")
        species = {sp.code: sp for sp in ses.exec(select(Species)).all()}
        arrivals = _cache(ses, LogArrival, "arrival_code")
        logs = _cache(ses, Log, "log_number")

        def get_species(name: str) -> Species | None:
            name = s(name)
            if not name:
                return None
            code = name.upper()
            sp = species.get(code)
            if not sp:
                sp = Species(code=code, common_name=name)
                ses.add(sp)
                ses.flush()
                species[code] = sp
                rep.bump("Species", "auto-created")
            return sp

        # ── WarehouseLocation ──
        if "WarehouseLocation" in wb.sheetnames:
            for r, d in sheet_rows(wb["WarehouseLocation"]):
                code = s(d.get("code"))
                if not code:
                    rep.error("WarehouseLocation", r, "missing code"); continue
                o = locs.get(code) or WarehouseLocation(code=code, name=code)
                o.name = s(d.get("name")) or o.name
                o.kind = s(d.get("kind"))
                o.notes = s(d.get("notes"))
                if code not in locs:
                    ses.add(o); locs[code] = o; rep.bump("WarehouseLocation", "created")
                else:
                    rep.bump("WarehouseLocation", "updated")

        # ── Supplier ──
        if "Supplier" in wb.sheetnames:
            for r, d in sheet_rows(wb["Supplier"]):
                code = s(d.get("code"))
                if not code:
                    rep.error("Supplier", r, "missing code"); continue
                o = sups.get(code) or Supplier(code=code, name=code)
                o.name = s(d.get("name")) or o.name
                o.contact_person = s(d.get("contact_person"))
                o.phone = s(d.get("phone")); o.email = s(d.get("email"))
                o.address = s(d.get("address")); o.notes = s(d.get("notes"))
                if code not in sups:
                    ses.add(o); sups[code] = o; rep.bump("Supplier", "created")
                else:
                    rep.bump("Supplier", "updated")

        # ── Customer (keyed by name) ──
        if "Customer" in wb.sheetnames:
            for r, d in sheet_rows(wb["Customer"]):
                name = s(d.get("name"))
                if not name:
                    rep.error("Customer", r, "missing name"); continue
                o = custs.get(name) or Customer(name=name)
                o.contact_person = s(d.get("contact_person")); o.phone = s(d.get("phone"))
                o.email = s(d.get("email")); o.address = s(d.get("address"))
                o.notes = s(d.get("notes"))
                if name not in custs:
                    ses.add(o); custs[name] = o; rep.bump("Customer", "created")
                else:
                    rep.bump("Customer", "updated")

        ses.flush()

        # ── Item ──
        if "Item" in wb.sheetnames:
            valid_kinds = {k.value for k in ItemKind}
            for r, d in sheet_rows(wb["Item"]):
                code = s(d.get("code"))
                if not code:
                    rep.error("Item", r, "missing code"); continue
                kind = s(d.get("kind")).upper()
                if kind not in valid_kinds:
                    rep.error("Item", r, f"bad kind '{kind}' (allowed: {sorted(valid_kinds)})")
                    continue
                sup_code = s(d.get("supplier_code"))
                sup = sups.get(sup_code) if sup_code else None
                if sup_code and not sup:
                    rep.error("Item", r, f"unknown supplier_code '{sup_code}'"); continue
                sp = get_species(d.get("species"))
                o = items.get(code) or Item(code=code, name=code, kind=ItemKind(kind), unit="")
                o.kind = ItemKind(kind)
                o.name = s(d.get("name")) or o.name
                o.name_th = s(d.get("name_th")); o.name_zh = s(d.get("name_zh"))
                o.unit = s(d.get("unit")) or o.unit
                o.species_id = sp.id if sp else None
                o.grade = s(d.get("grade")) or None
                o.cut_type = s(d.get("cut_type")) or None
                o.matching = s(d.get("matching")) or None
                o.face_back = s(d.get("face_back")) or None
                o.fsc = s(d.get("fsc")) or None
                o.board_type = s(d.get("board_type")) or None
                o.glue_type = s(d.get("glue_type")) or None
                o.thickness_mm = num(d.get("thickness_mm"))
                o.width_mm = num(d.get("width_mm")); o.length_mm = num(d.get("length_mm"))
                o.unit_cost = num(d.get("unit_cost")) or 0.0
                o.price = num(d.get("price")) or 0.0
                o.acc_code = s(d.get("acc_code")) or None
                o.reorder_point = num(d.get("reorder_point")) or 0.0
                o.supplier_id = sup.id if sup else None
                o.notes = s(d.get("notes"))
                if code not in items:
                    ses.add(o); items[code] = o; rep.bump("Item", "created")
                else:
                    rep.bump("Item", "updated")
            ses.flush()

        # ── GlueRecipe ──
        if "GlueRecipe" in wb.sheetnames:
            for r, d in sheet_rows(wb["GlueRecipe"]):
                code = s(d.get("recipe_code"))
                if not code:
                    rep.error("GlueRecipe", r, "missing recipe_code"); continue
                o = recipes.get(code) or GlueRecipe(recipe_code=code, name=code)
                o.name = s(d.get("name")) or o.name
                o.resin_ratio = num(d.get("resin_ratio")) or 0.0
                o.hardener_ratio = num(d.get("hardener_ratio")) or 0.0
                o.extender_ratio = num(d.get("extender_ratio")) or 0.0
                o.filler_ratio = num(d.get("filler_ratio")) or 0.0
                o.water_ratio = num(d.get("water_ratio")) or 0.0
                o.mix_time_min = intnum(d.get("mix_time_min")) or 0
                o.notes = s(d.get("notes"))
                if code not in recipes:
                    ses.add(o); recipes[code] = o; rep.bump("GlueRecipe", "created")
                else:
                    rep.bump("GlueRecipe", "updated")
            ses.flush()

        # ── GlueRecipe_Components -> recipe.material_links {role: item_id} ──
        if "GlueRecipe_Components" in wb.sheetnames:
            links: dict[str, dict] = {}
            for r, d in sheet_rows(wb["GlueRecipe_Components"]):
                rc = s(d.get("recipe_code")); role = s(d.get("component_role"))
                ic = s(d.get("item_code"))
                if not (rc and role and ic):
                    rep.error("GlueRecipe_Components", r, "recipe_code/component_role/item_code required")
                    continue
                if rc not in recipes:
                    rep.error("GlueRecipe_Components", r, f"unknown recipe_code '{rc}'"); continue
                if ic not in items:
                    rep.error("GlueRecipe_Components", r, f"unknown item_code '{ic}'"); continue
                links.setdefault(rc, {})[role] = {"item_id": items[ic].id,
                                                  "ratio": num(d.get("ratio"))}
                rep.bump("GlueRecipe_Components", "linked")
            for rc, m in links.items():
                recipes[rc].material_links = m   # reassign so JSONB change is tracked

        # ── LogArrival ──
        if "LogArrival" in wb.sheetnames:
            for r, d in sheet_rows(wb["LogArrival"]):
                code = s(d.get("arrival_code"))
                if not code:
                    rep.error("LogArrival", r, "missing arrival_code"); continue
                sup_code = s(d.get("supplier_code"))
                sup = sups.get(sup_code) if sup_code else None
                if sup_code and not sup:
                    rep.error("LogArrival", r, f"unknown supplier_code '{sup_code}'"); continue
                o = arrivals.get(code) or LogArrival(arrival_code=code)
                o.arrival_date = to_date(d.get("arrival_date"))
                o.supplier_id = sup.id if sup else None
                o.container_ref = s(d.get("container_ref")); o.notes = s(d.get("notes"))
                if code not in arrivals:
                    ses.add(o); arrivals[code] = o; rep.bump("LogArrival", "created")
                else:
                    rep.bump("LogArrival", "updated")
            ses.flush()

        # ── Log (with dual-unit fill) ──
        if "Log" in wb.sheetnames:
            for r, d in sheet_rows(wb["Log"]):
                ln = s(d.get("log_number"))
                if not ln:
                    rep.error("Log", r, "missing log_number"); continue
                ac = s(d.get("arrival_code"))
                arr = arrivals.get(ac)
                if not arr:
                    rep.error("Log", r, f"unknown arrival_code '{ac}'"); continue
                li, lm = num(d.get("length_in")), num(d.get("length_m"))
                di, dm = num(d.get("diameter_in")), num(d.get("diameter_m"))
                vf, vm = num(d.get("volume_ft3")), num(d.get("volume_m3"))
                if li and not lm: lm = li * IN_TO_M
                if lm and not li: li = lm / IN_TO_M
                if di and not dm: dm = di * IN_TO_M
                if dm and not di: di = dm / IN_TO_M
                if vf and not vm: vm = vf * FT3_TO_M3
                if vm and not vf: vf = vm / FT3_TO_M3
                sp = get_species(d.get("species"))
                o = logs.get(ln) or Log(log_number=ln, arrival_id=arr.id, log_code="")
                o.arrival_id = arr.id
                o.log_code = s(d.get("log_code")) or o.log_code
                o.species_id = sp.id if sp else None
                o.grade = s(d.get("grade")) or None
                o.length_in, o.length_m = li, lm
                o.diameter_in, o.diameter_m = di, dm
                o.volume_ft3, o.volume_m3 = vf, vm
                o.notes = s(d.get("notes"))
                if ln not in logs:
                    ses.add(o); logs[ln] = o; rep.bump("Log", "created")
                else:
                    rep.bump("Log", "updated")

        ses.commit()
    rep.print()


if __name__ == "__main__":
    run(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH)
