"""Import BOMs from templates/pvwood_bom.xlsx into Postgres.

- Assembly_BOM: one row per SKU -> upserts the Product (auto-masters
  product_category) + rebuilds its ASSEMBLY BOM lines.
- Transform_PVS / Transform_PSP -> upserts TransformationRecipe + lines.

Idempotent. Resolves item / glue-recipe / packing codes; reports unknowns.
Run master-data import FIRST. From repo root:  python -m scripts.import_bom [path]
"""
from __future__ import annotations

import sys

from openpyxl import load_workbook
from sqlmodel import Session, delete, select

from app.models.bom import BOMConsumable, BOMHeader, BOMItem, GlueRecipe
from app.models.enums import BOMLineRole, BOMType, ItemKind
from app.models.master import Item, Product, ProductCategory
from app.models.transformation_recipe import TransformationRecipe, TransformationRecipeLine
from app.db import engine
from scripts._import_utils import Report, intnum, num, s, sheet_rows

DEFAULT_PATH = "templates/pvwood_bom.xlsx"


def _cache(ses: Session, model, key: str) -> dict:
    return {getattr(o, key): o for o in ses.exec(select(model)).all()}


def run(path: str) -> None:
    wb = load_workbook(path, data_only=True)
    rep = Report()
    with Session(engine) as ses:
        items = _cache(ses, Item, "code")
        recipes = _cache(ses, GlueRecipe, "recipe_code")
        products = _cache(ses, Product, "code")
        cats = {c.name: c for c in ses.exec(select(ProductCategory)).all()}

        def get_category(name: str) -> ProductCategory | None:
            name = s(name)
            if not name:
                return None
            c = cats.get(name)
            if not c:
                c = ProductCategory(code=name.upper().replace(" ", "_")[:32], name=name)
                ses.add(c); ses.flush(); cats[name] = c
                rep.bump("ProductCategory", "auto-created")
            return c

        # (Glue recipes are imported from the glue CSV export via import_glue_csv;
        #  here we only RESOLVE face/back_glue_code against them.)

        # ── Assembly_BOM ──
        if "Assembly_BOM" in wb.sheetnames:
            for r, d in sheet_rows(wb["Assembly_BOM"]):
                sku = s(d.get("product_sku"))
                if not sku:
                    rep.error("Assembly_BOM", r, "missing product_sku"); continue

                # resolve every referenced code first; skip the whole SKU on any miss
                def ref_item(col):
                    code = s(d.get(col))
                    if not code:
                        return None, True
                    it = items.get(code)
                    if not it:
                        rep.error("Assembly_BOM", r, f"unknown item code '{code}' ({col})")
                        return None, False
                    return it, True

                def ref_recipe(col):
                    code = s(d.get(col))
                    if not code:
                        return None, True
                    gr = recipes.get(code)
                    if not gr:
                        rep.error("Assembly_BOM", r, f"unknown glue recipe '{code}' ({col})")
                        return None, False
                    return gr, True

                board, ok1 = ref_item("base_board_code")
                fv, ok2 = ref_item("face_veneer_code")
                bv, ok3 = ref_item("back_veneer_code")
                fc, ok4 = ref_item("face_coating_code")
                bc, ok5 = ref_item("back_coating_code")
                pk, ok6 = ref_item("packing_sku_code")
                fg, ok7 = ref_recipe("face_glue_code")
                bg, ok8 = ref_recipe("back_glue_code")
                if not all([ok1, ok2, ok3, ok4, ok5, ok6, ok7, ok8]):
                    continue

                cat = get_category(d.get("product_category"))
                p = products.get(sku) or Product(code=sku, name=s(d.get("sku_name")) or sku)
                p.name = s(d.get("sku_name")) or p.name
                p.category_id = cat.id if cat else None
                p.thickness_mm = num(d.get("thickness_mm"))
                p.width_mm = num(d.get("width_mm")); p.length_mm = num(d.get("length_mm"))
                p.pallet_qty = intnum(d.get("pieces_per_unit")) or 1
                p.notes = s(d.get("Notes"))
                if sku not in products:
                    ses.add(p); products[sku] = p; rep.bump("Product", "created")
                else:
                    rep.bump("Product", "updated")
                ses.flush()

                hdr = ses.exec(select(BOMHeader).where(
                    BOMHeader.product_id == p.id, BOMHeader.revision == 0)).first()
                if not hdr:
                    hdr = BOMHeader(product_id=p.id, bom_type=BOMType.ASSEMBLY)
                    ses.add(hdr); ses.flush(); rep.bump("BOMHeader", "created")
                else:
                    rep.bump("BOMHeader", "updated")
                # rebuild lines (idempotent)
                ses.exec(delete(BOMItem).where(BOMItem.bom_header_id == hdr.id))
                ses.exec(delete(BOMConsumable).where(BOMConsumable.bom_header_id == hdr.id))

                seq = 0
                lines = 0
                for it, qty_col, role in [(board, "base_board_qty", BOMLineRole.BOARD),
                                          (fv, "face_veneer_qty", BOMLineRole.FACE_VENEER),
                                          (bv, "back_veneer_qty", BOMLineRole.BACK_VENEER)]:
                    if it:
                        seq += 1
                        ses.add(BOMItem(bom_header_id=hdr.id, item_id=it.id, seq=seq,
                                        role=role, qty_override=num(d.get(qty_col))))
                        lines += 1
                for gr, use_col, role in [(fg, "face_glue_qty", BOMLineRole.FACE_GLUE),
                                          (bg, "back_glue_qty", BOMLineRole.BACK_GLUE)]:
                    if gr:
                        seq += 1
                        ses.add(BOMConsumable(bom_header_id=hdr.id, glue_recipe_id=gr.id,
                                              seq=seq, role=role,
                                              usage_g_per_face=num(d.get(use_col))))
                        lines += 1
                for it, use_col, side in [(fc, "Face_Coating_g/M2", "face"),
                                          (bc, "Back_Coating_g/M2", "back")]:
                    if it:
                        seq += 1
                        ses.add(BOMConsumable(bom_header_id=hdr.id, item_id=it.id, seq=seq,
                                              role=BOMLineRole.COATING,
                                              usage_g_per_face=num(d.get(use_col)),
                                              notes=side))
                        lines += 1
                if pk:
                    seq += 1
                    ses.add(BOMConsumable(bom_header_id=hdr.id, item_id=pk.id, seq=seq,
                                          role=BOMLineRole.PACKING, qty=1))
                    lines += 1
                rep.bump("BOM lines", "written", lines)
            ses.flush()

        # ── DoorSkin_BOM (PUV Mode-A: board + sealer + 2 primer passes + packing) ──
        if "DoorSkin_BOM" in wb.sheetnames:
            for r, d in sheet_rows(wb["DoorSkin_BOM"]):
                sku = s(d.get("product_sku"))
                if not sku:
                    rep.error("DoorSkin_BOM", r, "missing product_sku"); continue

                def ds_ref(col):
                    code = s(d.get(col))
                    if not code:
                        return None, True
                    it = items.get(code)
                    if not it:
                        rep.error("DoorSkin_BOM", r, f"unknown item code '{code}' ({col})")
                        return None, False
                    return it, True

                board, o1 = ds_ref("base_board_code")
                sealer, o2 = ds_ref("sealer_code")
                p1, o3 = ds_ref("primer1_code")
                p2, o4 = ds_ref("primer2_code")
                pk, o5 = ds_ref("packing_sku_code")
                if not all([o1, o2, o3, o4, o5]):
                    continue

                cat = get_category("Door Skin")
                p = products.get(sku) or Product(code=sku, name=s(d.get("sku_name")) or sku)
                p.name = s(d.get("sku_name")) or p.name
                p.category_id = cat.id if cat else None
                p.thickness_mm = num(d.get("thickness_mm"))
                p.width_mm = num(d.get("width_mm")); p.length_mm = num(d.get("length_mm"))
                p.pallet_qty = intnum(d.get("pieces_per_unit")) or 1
                p.notes = s(d.get("Notes"))
                if sku not in products:
                    ses.add(p); products[sku] = p; rep.bump("DoorSkin Product", "created")
                else:
                    rep.bump("DoorSkin Product", "updated")
                ses.flush()

                hdr = ses.exec(select(BOMHeader).where(
                    BOMHeader.product_id == p.id, BOMHeader.revision == 0)).first()
                if not hdr:
                    hdr = BOMHeader(product_id=p.id, bom_type=BOMType.ASSEMBLY)
                    ses.add(hdr); ses.flush()
                ses.exec(delete(BOMItem).where(BOMItem.bom_header_id == hdr.id))
                ses.exec(delete(BOMConsumable).where(BOMConsumable.bom_header_id == hdr.id))

                seq = 0
                if board:
                    seq += 1
                    ses.add(BOMItem(bom_header_id=hdr.id, item_id=board.id, seq=seq,
                                    role=BOMLineRole.BOARD, qty_override=num(d.get("base_board_qty"))))
                for it, gcol, role, note in [
                        (sealer, "sealer_g_m2", BOMLineRole.COATING, "sealer"),
                        (p1, "primer1_g_m2", BOMLineRole.PRIMER, "pass1"),
                        (p2, "primer2_g_m2", BOMLineRole.PRIMER, "pass2")]:
                    if it:
                        seq += 1
                        ses.add(BOMConsumable(bom_header_id=hdr.id, item_id=it.id, seq=seq,
                                              role=role, usage_g_per_face=num(d.get(gcol)), notes=note))
                if pk:
                    seq += 1
                    ses.add(BOMConsumable(bom_header_id=hdr.id, item_id=pk.id, seq=seq,
                                          role=BOMLineRole.PACKING, qty=1))
                rep.bump("DoorSkin lines", "written", seq)
            ses.flush()

        # ── Transform_PVS / Transform_PSP ──
        for sheet, line_code in [("Transform_PVS", "PVS"), ("Transform_PSP", "PSP")]:
            if sheet not in wb.sheetnames:
                continue
            rows = list(sheet_rows(wb[sheet]))
            by_recipe: dict[str, list] = {}
            for r, d in rows:
                rc = s(d.get("recipe_code"))
                if not rc:
                    rep.error(sheet, r, "missing recipe_code"); continue
                by_recipe.setdefault(rc, []).append((r, d))
            for rc, rlist in by_recipe.items():
                name = next((s(d.get("name")) for _, d in rlist if s(d.get("name"))), rc)
                rec = ses.exec(select(TransformationRecipe).where(
                    TransformationRecipe.code == rc)).first()
                if not rec:
                    rec = TransformationRecipe(code=rc, name=name, line_code=line_code)
                    ses.add(rec); ses.flush(); rep.bump(sheet, "recipe created")
                else:
                    rec.name = name; rec.line_code = line_code; rep.bump(sheet, "recipe updated")
                ses.exec(delete(TransformationRecipeLine).where(
                    TransformationRecipeLine.recipe_id == rec.id))
                for r, d in rlist:
                    ic = s(d.get("item_code"))
                    kind = s(d.get("kind")).upper() or "OUTPUT"
                    it = items.get(ic) if ic else None
                    if ic and not it:
                        if kind == "OUTPUT":   # produced (sliced/spliced) veneer -> create item
                            it = Item(
                                code=ic, kind=ItemKind.VENEER, name=ic,
                                unit=s(d.get("uom")) or "",
                                grade=s(d.get("grade")) or None,
                                cut_type=s(d.get("cut")) or None,
                                matching=s(d.get("mat")) or None,
                                fsc=s(d.get("FSC")) or s(d.get("fsc")) or None,
                                thickness_mm=num(d.get("Thickness_mm")) or num(d.get("thickness_mm")),
                                width_mm=num(d.get("width_mm")) or num(d.get("Width_mm")),
                                length_mm=num(d.get("length_mm")) or num(d.get("Length_mm")))
                            ses.add(it); ses.flush(); items[ic] = it
                            rep.bump(sheet, "produced-veneer+")
                        else:
                            rep.error(sheet, r, f"unknown item_code '{ic}'"); continue
                    ses.add(TransformationRecipeLine(
                        recipe_id=rec.id, kind=kind,
                        seq=intnum(d.get("seq")) or 0,
                        item_id=it.id if it else None, role=s(d.get("role")) or None,
                        qty=num(d.get("qty")), uom=s(d.get("uom")) or None,
                        grade=s(d.get("grade")) or None,
                        expected_yield_pct=num(d.get("expected_yield_pct")),
                        expected_loss_pct=num(d.get("expected_loss_pct")),
                        notes=s(d.get("notes"))))
                    rep.bump(sheet, "lines")

        ses.commit()
    rep.print()


if __name__ == "__main__":
    run(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH)
