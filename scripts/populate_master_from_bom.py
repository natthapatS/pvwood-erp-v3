"""Pre-populate a master-data workbook's Item sheet from the owner's BOM workbook.

Pulls every distinct item code the BOMs/recipes reference:
  - VENEER items from the V + M2 sheets (with species/dims/grade/cut/…)
  - BOARD codes from Assembly_BOM.base_board_code
  - LOG codes from Transform_PVS inputs
  - PACKING codes from Assembly_BOM.packing_sku_code
  - any veneer referenced but missing from V/M2 (added blank, kind=VENEER)

Writes a master-data workbook so only descriptions/costs/supplier need filling in.
Run from repo root:
    python -m scripts.populate_master_from_bom <bom.xlsx> [out.xlsx]
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.build_templates import MASTER_SHEETS, _add_sheet, _instructions


def S(v) -> str:
    return "" if v is None else str(v).strip()


def numS(v):
    v = S(v)
    try:
        return float(v)
    except ValueError:
        return ""


def _hdr_map(ws):
    """English prefix of each header (strips the Chinese line) -> column index."""
    m = {}
    for i, c in enumerate(ws[1]):
        h = S(c.value).split("\n")[0].strip().lower()
        if h and h not in m:
            m[h] = i
    return m


def _veneers(wb) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for sh in ("V", "M2"):
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        m = _hdr_map(ws)

        def col(row, *names):
            for n in names:
                i = m.get(n)
                if i is not None and i < len(row):
                    return row[i]
            return None

        for row in ws.iter_rows(min_row=2, values_only=True):
            code = S(col(row, "code"))
            if not code:
                continue
            out[code] = {
                "code": code, "kind": "VENEER",
                "name": S(col(row, "name")),
                "unit": S(col(row, "unit")),
                "species": S(col(row, "species")),
                "grade": S(col(row, "grade")),
                "cut_type": S(col(row, "cut")),
                "matching": S(col(row, "mat.", "mat")),
                "face_back": S(col(row, "f/b")),
                "thickness_mm": numS(col(row, "v-thi", "thickness")),
                "width_mm": numS(col(row, "width")),
                "length_mm": numS(col(row, "length")),
            }
    return out


def _asm_refs(wb) -> tuple[set, set, set, set]:
    boards, veneers, packing = set(), set(), set()
    logs: set = set()
    if "Assembly_BOM" in wb.sheetnames:
        ws = wb["Assembly_BOM"]
        hdr = [S(c.value).replace("★", "").strip() for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {h: row[i] for i, h in enumerate(hdr) if h and i < len(row)}
            if S(d.get("base_board_code")):
                boards.add(S(d["base_board_code"]))
            for k in ("face_veneer_code", "back_veneer_code"):
                if S(d.get(k)):
                    veneers.add(S(d[k]))
            if S(d.get("packing_sku_code")):
                packing.add(S(d["packing_sku_code"]))
    if "Transform_PVS" in wb.sheetnames:
        ws = wb["Transform_PVS"]
        hdr = [S(c.value).replace("★", "").strip() for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {h: row[i] for i, h in enumerate(hdr) if h and i < len(row)}
            if S(d.get("kind")).upper() == "INPUT" and S(d.get("item_code")):
                logs.add(S(d["item_code"]))
    return boards, veneers, packing, logs


def build(bom_path: str, out_path: str) -> None:
    wb = load_workbook(bom_path, data_only=True)
    veneers = _veneers(wb)
    boards, ref_veneers, packing, logs = _asm_refs(wb)

    items: dict[str, dict] = dict(veneers)
    for c in ref_veneers:                      # veneers used but missing from V/M2
        items.setdefault(c, {"code": c, "kind": "VENEER"})
    for c in boards:
        items.setdefault(c, {"code": c, "kind": "BOARD"})
    for c in logs:
        items.setdefault(c, {"code": c, "kind": "LOG"})
    for c in packing:
        items.setdefault(c, {"code": c, "kind": "PACKING"})

    # ── write workbook ──
    out = Workbook()
    out.remove(out.active)
    for name, (cols, eg) in MASTER_SHEETS.items():
        _add_sheet(out, name, cols, eg)
    _instructions(out, "PVWood ERP v3 — Master Data (pre-populated Item codes)", [
        "The Item sheet is pre-filled with every code your BOMs reference.",
        "Veneers carry species/dims/grade/cut/matching from your V + M2 sheets.",
        "Boards, logs and packing are listed as codes only — fill name/dims/cost/etc.",
        "Add name_th/name_zh, unit_cost, price, acc_code, reorder_point, supplier_code as known.",
        "Fill WarehouseLocation + Supplier + Customer on their sheets.",
    ])

    ws = out["Item"]
    cols = [name for (name, _req, _dd) in MASTER_SHEETS["Item"][0]]
    for code in sorted(items):
        d = items[code]
        ws.append([d.get(c, "") for c in cols])

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    out.save(out_path)
    kinds: dict[str, int] = {}
    for d in items.values():
        kinds[d["kind"]] = kinds.get(d["kind"], 0) + 1
    print(f"wrote {out_path}: {len(items)} items ->", kinds)


if __name__ == "__main__":
    bom = sys.argv[1] if len(sys.argv) > 1 else "templates/pvwood_bom_20260714.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "templates/pvwood_master_data_prefilled.xlsx"
    build(bom, out)
