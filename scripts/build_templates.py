"""Generate the blank data-entry templates (master data + BOMs).

Reproducible: the column specs live here; the .xlsx files under templates/ are
generated output. Run from repo root:  python -m scripts.build_templates

Each sheet: bold header row (required columns marked ★ + shaded), one example row,
enum dropdowns, frozen header. An Instructions sheet leads each workbook.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.models.enums import BOMLineRole, CalcMethod, ItemKind

TEMPLATES_DIR = Path(__file__).resolve().parent.parent / "templates"

# Enum value lists (single source = the model enums where they exist).

ITEM_KINDS = [e.value for e in ItemKind]
BOM_ROLES = [e.value for e in BOMLineRole]
CALC_METHODS = [e.value for e in CalcMethod]
LOCATION_KINDS = ["WAREHOUSE", "FC", "WIP", "FG", "STAGING"]
RECIPE_LINE_KINDS = ["INPUT", "OUTPUT", "STEP"]

REQ_FILL = PatternFill("solid", fgColor="FCE4E4")   # light red = required
HDR_FONT = Font(bold=True)
EG_FONT = Font(italic=True, color="888888")

# sheet -> (columns, example_row). A column is (name, required, dropdown|None).
MASTER_SHEETS: dict[str, tuple[list[tuple[str, bool, list | None]], list]] = {
    "Species": (
        [("code", True, None), ("common_name", True, None),
         ("botanical_name", False, None), ("notes", False, None)],
        ["OAK", "White Oak", "Quercus alba", ""],
    ),
    "ProductCategory": (
        [("code", True, None), ("name", True, None), ("description", False, None)],
        ["DOORSKIN", "Door Skin", "Primed/UV door skins"],
    ),
    "WarehouseLocation": (
        [("code", True, None), ("name", True, None),
         ("kind", False, LOCATION_KINDS), ("notes", False, None)],
        ["WH-A1", "Warehouse A Rack 1", "WAREHOUSE", ""],
    ),
    "Supplier": (
        [("code", True, None), ("name", True, None), ("contact_person", False, None),
         ("phone", False, None), ("email", False, None), ("address", False, None),
         ("notes", False, None)],
        ["SUP001", "ACME Veneer Co", "K. Somchai", "0812345678", "", "", ""],
    ),
    "Customer": (
        [("name", True, None), ("contact_person", False, None), ("phone", False, None),
         ("email", False, None), ("address", False, None), ("notes", False, None)],
        ["Global Doors Ltd", "J. Smith", "", "", "", ""],
    ),
    "Item": (
        [("code", True, None), ("kind", True, ITEM_KINDS), ("name", True, None),
         ("name_th", False, None), ("name_zh", False, None), ("unit", True, None),
         ("species_code", False, None), ("grade", False, None), ("cut_type", False, None),
         ("matching", False, None), ("face_back", False, None), ("fsc", False, None),
         ("board_type", False, None), ("glue_type", False, None),
         ("thickness_mm", False, None), ("width_mm", False, None), ("length_mm", False, None),
         ("unit_cost", False, None), ("price", False, None), ("acc_code", False, None),
         ("reorder_point", False, None), ("supplier_code", False, None), ("notes", False, None)],
        ["VNR-OAK-CQ-A", "VENEER", "Oak Crown A", "", "", "m2", "OAK", "A", "CROWN",
         "BOOK", "FACE", "FSC100", "", "", "0.5", "1270", "2500", "45", "0", "5201",
         "1000", "SUP001", ""],
    ),
    "GlueRecipe": (
        [("recipe_code", True, None), ("name", True, None), ("resin_ratio", False, None),
         ("hardener_ratio", False, None), ("extender_ratio", False, None),
         ("filler_ratio", False, None), ("water_ratio", False, None),
         ("mix_time_min", False, None), ("notes", False, None)],
        ["GL-STD", "Standard UF", "100", "20", "10", "5", "0", "20", ""],
    ),
    "GlueRecipe_Components": (
        [("recipe_code", True, None), ("component_role", True, None),
         ("item_code", True, None), ("ratio", False, None)],
        ["GL-STD", "resin", "GLU-UF-RESIN", "100"],
    ),
    "LogArrival": (
        [("arrival_code", True, None), ("arrival_date", False, None),
         ("supplier_code", False, None), ("container_ref", False, None), ("notes", False, None)],
        ["ARR-2026-001", "2026-07-01", "SUP001", "MSKU1234567", ""],
    ),
    "Log": (
        [("log_number", True, None), ("arrival_code", True, None), ("log_code", True, None),
         ("species_code", False, None), ("grade", False, None),
         ("length_in", False, None), ("length_m", False, None),
         ("diameter_in", False, None), ("diameter_m", False, None),
         ("volume_ft3", False, None), ("volume_m3", False, None), ("notes", False, None)],
        ["LOG-0001", "ARR-2026-001", "OAK-A", "OAK", "A", "98", "2.5", "24", "0.61",
         "", "1.2", "photo in notes"],
    ),
}

BOM_SHEETS: dict[str, tuple[list[tuple[str, bool, list | None]], list]] = {
    # Linear (one row per SKU) assembly BOM — owner's column layout. Also defines
    # the product (sku_name / pieces_per_unit / dims), so there is no separate
    # Product sheet: the importer upserts the Product then its BOM lines.
    "Assembly_BOM": (
        [("product_sku", True, None), ("sku_name", False, None),
         ("pieces_per_unit", False, None), ("thickness_mm", False, None),
         ("width_mm", False, None), ("length_mm", False, None),
         ("base_board_code", False, None), ("base_board_qty", False, None),
         ("face_veneer_code", False, None), ("face_veneer_qty", False, None),
         ("face_glue_code", False, None), ("face_glue_qty", False, None),
         ("back_veneer_code", False, None), ("back_veneer_qty", False, None),
         ("back_glue_code", False, None), ("back_glue_qty", False, None),
         ("face_coating_code", False, None), ("Face_Coating_g/M2", False, None),
         ("back_coating_code", False, None), ("Back_Coating_g/M2", False, None),
         ("packing_sku_code", False, None), ("Notes", False, None)],
        ["PNL-OAK", "Oak Overlay Panel", "300", "3.2", "915", "2135",
         "BRD-MDF-3.2", "1", "VNR-OAK-A", "1", "GL-STD", "120",
         "VNR-OAK-B", "1", "GL-STD", "120", "", "", "", "", "PKG-CARTON", ""],
    ),
    "Transform_PVS": (
        [("recipe_code", True, None), ("name", True, None), ("kind", True, RECIPE_LINE_KINDS),
         ("seq", False, None), ("item_code", False, None), ("role", False, None),
         ("qty", False, None), ("uom", False, None), ("grade", False, None),
         ("expected_yield_pct", False, None), ("expected_loss_pct", False, None),
         ("notes", False, None)],
        ["PVS-OAK", "Oak slicing", "INPUT", "1", "OAK-A", "log", "1", "m3", "A", "", "", ""],
    ),
    "Transform_PSP": (
        [("recipe_code", True, None), ("name", True, None), ("kind", True, RECIPE_LINE_KINDS),
         ("seq", False, None), ("item_code", False, None), ("role", False, None),
         ("qty", False, None), ("uom", False, None), ("grade", False, None),
         ("expected_yield_pct", False, None), ("expected_loss_pct", False, None),
         ("notes", False, None)],
        ["PSP-OAK", "Oak splicing", "INPUT", "1", "VNR-OAK-CQ-A", "veneer", "1", "m2", "",
         "", "", ""],
    ),
}


def _add_sheet(wb: Workbook, name: str, columns, example) -> None:
    ws = wb.create_sheet(name)
    for c, (col, required, dropdown) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c, value=("★ " if required else "") + col)
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="left")
        if required:
            cell.fill = REQ_FILL
        ws.column_dimensions[cell.column_letter].width = max(12, len(col) + 4)
        if example and c <= len(example):
            ws.cell(row=2, column=c, value=example[c - 1]).font = EG_FONT
        if dropdown:
            dv = DataValidation(type="list", formula1='"' + ",".join(dropdown) + '"',
                                allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{cell.column_letter}3:{cell.column_letter}2000")
    ws.freeze_panes = "A2"


def _instructions(wb: Workbook, title: str, lines: list[str]) -> None:
    ws = wb.create_sheet("Instructions", 0)
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    for i, line in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100


def build_master() -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    for name, (cols, eg) in MASTER_SHEETS.items():
        _add_sheet(wb, name, cols, eg)
    _instructions(wb, "PVWood ERP v3 — Master Data", [
        "★ = required column (shaded). Row 2 (grey italic) is an example — overwrite or delete it.",
        "Codes are your own identifiers; other sheets/BOMs reference rows by these codes.",
        "Fill reference sheets first (Species, ProductCategory, WarehouseLocation, Supplier),",
        "then Item / GlueRecipe, then GlueRecipe_Components, LogArrival, Log.",
        "Products (finished SKUs) are defined in the BOM workbook's Assembly_BOM sheet, not here.",
        "Item.kind and Location.kind use dropdowns. Dimensions in mm. Costs per unit.",
        "Logs: enter any one unit of length/diameter/volume — the system converts the rest.",
    ])
    out = TEMPLATES_DIR / "pvwood_master_data.xlsx"
    TEMPLATES_DIR.mkdir(exist_ok=True)
    wb.save(out)
    return out


def build_bom() -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    for name, (cols, eg) in BOM_SHEETS.items():
        _add_sheet(wb, name, cols, eg)
    _instructions(wb, "PVWood ERP v3 — BOMs", [
        "★ = required. Fill master data first (items / glue recipes must already exist).",
        "ONE ROW PER SKU. The row defines the product (sku_name, pieces_per_unit, dims) AND",
        "its recipe. Fill each component's code + qty; leave a cell blank if unused (e.g. no",
        "back veneer). Glue can differ face vs back (face_glue_code / back_glue_code).",
        "Coating (face/back _coating_code + g/M2) = a primer (door skins) OR a UV topcoat",
        "  (panels). Fill only the side(s) that get coated — e.g. a door skin = base board +",
        "  face_coating_code + packing; a UV-topcoat panel = its normal rows + a coating.",
        "Transform_PVS / Transform_PSP: recipe lines (kind = INPUT/OUTPUT/STEP), one row each.",
        "All *_code / *_sku cells reference codes you defined in the master-data workbook.",
    ])
    out = TEMPLATES_DIR / "pvwood_bom.xlsx"
    TEMPLATES_DIR.mkdir(exist_ok=True)
    wb.save(out)
    return out


if __name__ == "__main__":
    print("wrote", build_master())
    print("wrote", build_bom())
